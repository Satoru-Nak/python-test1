import os
import datetime
import re
import telmod
import glob
import openpyxl as px

log_file_base_path = r"C:\Users\NTTCOM\SoftwareEng\python-test1"

def create_log_file(router_name , test_id):
    """
    フォルダ："base_path + test_id"フォルダに保存 test_idにrollbackが含まれていた場合削除
    ファイル："test_id + router_name + 時間"ファイルを作成→保存
    """
    nowstr = str(datetime.datetime.now().strftime("%Y%m%d-%H%M(%S)"))
    #logファイル名
    logfilename = test_id + "-" + router_name + "-" + nowstr + "-log.txt"

    #保存場所　rollbackの場合も同じフォルダに保存（ファイル名は↑で別にする）
    fullfilepath = log_file_base_path + "\\" + re.sub(r"_rollback" , "" , test_id)
    #print("DEBUG log file created at " + fullfilepath + "\\" + logfilename)

    #create folder if it doesn't exist
    if not os.path.exists(fullfilepath):
        os.mkdir(fullfilepath)
        print("created files for " + test_id)

    return open(fullfilepath + "\\" + logfilename, "w")

def init_telnet_proc(r_list, test_router_list, test_id):
    """
    r_listからtelnetセッションを生成
    引数：r_list : router情報dict, test_router_list : testに必要なrouter名list
    戻り：RemoteAccessインスタンスをdictにまとめて、ログ用ファイルハンドラをdictにまとめて
    処理概要：
    1. r_listにtestに必要なルータ情報が含まれているかをチェック
    2. 初期処理、ログファイル生成、RemoteAccessクラスインスタンス化、telnet（ログインまで）
    """
    raw_fn = {}
    telctl = {}
    error_flg = False
    #r_listにtestに必要なルータ情報が含まれているかをチェック
    for router_name in test_router_list:
        if router_name not in r_list:
            error_flg = True
            break

    if error_flg:
        print("r_list not defined correctly...")
    else:
        #初期準備  r_listからlog fileを生成＋初期接続
        for r_name,r_data in r_list.items():
            raw_fn[r_name] = create_log_file(r_name,test_id)
            telctl[r_name] = telmod.RemoteAccess(**r_data)
            telctl[r_name].init_telnet()

        return([telctl , raw_fn])

def send_com_para(cmd_list, telctl, raw_fn):
    """
    cmd_listはkey=router名（r_listのkey),value=流し込むコマンドのリスト
    送られる順番はkeyのアルファベット順で、１つのルータですべてのコマンド実行が
    完了してから次のコマンド実行に移る
    """
    for rs, c_list in sorted(cmd_list.items()):
        #rsがtelctl, raw_fnに含まれることを確認して
        if rs in telctl and rs in raw_fn:
            #流し込み
            telctl[rs].send_com_list_wf(c_list, raw_fn[rs])
            telctl[rs].send_com(" ",raw_fn[rs])
        else:
            print("router not defined in cmd_list key...")

def terminate_telnet_proc(telctl, raw_fn):
    """
    終了処理　file close + telnet切断
    """
    for fn in raw_fn.values():
        fn.close()

    for telcl in telctl.values():
        telcl.disconnect()

def get_test_data():
    """
    test0001.xlsxから試験に必要なデータを読み込む
    現状　C2:test_id, C3:test_router_list
    A列に項番記述（6行目からint型にすること）B列に試験手順名
    C列にコマンドのリスト　D列にコマンド投入対象ルータ
    ★B列は項番記述行以外に何もいれないこと（Noneであること）
    ★完了の場合、A列に"END"文字列を記述すること

    返り値：test_id:string型、test_router_list:リスト型　test_proc:ネストされたリスト型
    test_procは[項番１のdata(リスト型), 項番２のdata(リスト型)...]
    項番xのdata(リスト型)は[投入router_list(リスト型), command 1(string型), command2...]
    """
    editbook = "test0001/test0001.xlsx"
    editsheet = "test_procedure"

    book = px.load_workbook(editbook) #book読み込み
    ws = book.worksheets[book.get_sheet_names().index(editsheet)] # sheet読み込み

    test_id = ws.cell(row=2,column=3).value#test-id読み込み
    test_router_list = ws.cell(row=3,column=3).value.split(",")#test_router_list読み込み

    i = 6
    proc_id = 0
    test_proc = []

    while True:
        #print("DEBUG: current i:{0}".format(i))
        if type(ws.cell(row=i,column=1).value) == int:#項番列が数値だったら
            proc_id += 1
            proc_tmp = []
            proc_tmp.append(ws.cell(row=i,column=4).value.split(","))#投入ルータリスト
            proc_tmp.append(r"!:" + ws.cell(row=i,column=2).value)#試験手順のコメント取得
            proc_tmp.append(ws.cell(row=i,column=3).value)#コメントの横のコマンド読み込み

            i += 1
            while ws.cell(row=i,column=2).value == None and ws.cell(row=i,column=1).value != "END":
                proc_tmp.append(ws.cell(row=i,column=3).value)
                i += 1

            test_proc.append(proc_tmp)

        elif ws.cell(row=i,column=1).value == "END":
            print("test data fetched. end reached with i = {0}".format(i))
            break
        else:
            print("invalid error in importing test data ...")
            break

    test_data = [test_id, test_router_list, test_proc]
    #print(test_data)
    return test_data
